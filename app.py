# -*- coding: utf-8 -*-
import datetime
import os
import sqlite3
import shutil
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image

DB_NAME = "materials_advanced.db"

# [1. Streamlit 웹 페이지 기본 설정]
st.set_page_config(
    page_title="(주)진테크 스마트 부자재 통합 재고 관리 솔루션",
    page_icon="🏬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- 데이터베이스 및 백업 로직 ---
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS items (
            name TEXT PRIMARY KEY,
            specification TEXT,
            safety_stock INTEGER DEFAULT 0,
            price INTEGER DEFAULT 0
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS workers (
            name TEXT PRIMARY KEY
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            name TEXT,
            division TEXT,
            qty INTEGER,
            manager TEXT,
            note TEXT
        )
    """)
    conn.commit()
    conn.close()
    
    try:
        today_str = datetime.date.today().strftime("%Y%m%d")
        backup_name = f"backup_{today_str}.db"
        if not os.path.exists(backup_name):
            shutil.copyfile(DB_NAME, backup_name)
    except Exception:
        pass

# --- 엑셀 서식 적용 공통 함수 ---
def apply_log_sheet_style(ws, h_fill, h_font, is_integrated=False):
    d_font = Font(name="맑은 고딕", size=10)
    t_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    for cell in ws[1]:
        cell.fill = h_fill; cell.font = h_font; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = t_border
    
    max_row = ws.max_row
    data_max_row = max_row - 1 if is_integrated else max_row
    
    for row in ws.iter_rows(min_row=2, max_row=data_max_row):
        for cell in row:
            cell.font = d_font; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = t_border
            if cell.column in [3, 4, 5] and isinstance(cell.value, (int, float)): 
                cell.number_format = "#,##0"

# --- 월간 통합 엑셀 바이너리 생성 함수 (웹 다운로드용) ---
def generate_monthly_excel(s_date, e_date):
    conn = sqlite3.connect(DB_NAME)
    query_out = f"""
        SELECT h.date AS '출고일자', h.name AS '품목명', h.qty AS '출고수량', i.price AS '단가', 
               (h.qty * i.price) AS '출고금액', h.manager AS '작업자(가져간사람)', h.note AS '비고' 
        FROM history h LEFT JOIN items i ON h.name = i.name
        WHERE h.division='출고' AND h.date BETWEEN '{s_date}' AND '{e_date}'
        ORDER BY h.date ASC, h.id ASC
    """
    df_out = pd.read_sql_query(query_out, conn)

    query_in = f"""
        SELECT h.date AS '입고일자', h.name AS '품목명', h.qty AS '입고수량', i.price AS '단가', 
               (h.qty * i.price) AS '입고금액', h.manager AS '입고처(발주업체)', h.note AS '비고' 
        FROM history h LEFT JOIN items i ON h.name = i.name
        WHERE h.division='입고' AND h.date BETWEEN '{s_date}' AND '{e_date}'
        ORDER BY h.date ASC, h.id ASC
    """
    df_in = pd.read_sql_query(query_in, conn)
    conn.close()

    if df_out.empty and df_in.empty:
        return None

    file_name = f"진테크_부자재_종합결산보고서_({s_date}_to_{e_date}).xlsx"
    wb = Workbook()
    ws_default = wb.active

    t_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    sum_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="A6A6A6"), bottom=Side(style="double", color="A6A6A6"))
    d_font = Font(name="맑은 고딕", size=10)
    sum_font = Font(name="맑은 고딕", size=10, bold=True, color="000000")
    sum_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 1. 출고 대장
    ws_out = wb.create_sheet(title="출고 대장")
    ws_out.append(["출고일자", "품목명", "출고수량", "단가", "출고금액", "작업자(가져간사람)", "비고"])
    for r in df_out.values.tolist(): ws_out.append(r)
    out_max = ws_out.max_row
    if out_max > 1:
        ws_out.append(["합계", "", f"=SUM(C2:C{out_max})", "", f"=SUM(E2:E{out_max})", "", ""])
        for cell in ws_out[out_max + 1]:
            cell.font = sum_font; cell.fill = sum_fill; cell.border = sum_border; cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column in [3, 5]: cell.number_format = "#,##0"
    fill_out = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
    apply_log_sheet_style(ws_out, fill_out, Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF"), is_integrated=True)
    for col in ws_out.columns:
        max_len = max(sum(2 if ord(c) > 128 else 1 for c in str(cell.value or '')) for cell in col)
        ws_out.column_dimensions[col[0].column_letter].width = max(max_len + 4, 16)

    # 2. 입고 대장
    ws_in = wb.create_sheet(title="입고 대장")
    ws_in.append(["입고일자", "품목명", "입고수량", "단가", "입고금액", "입고처(발주업체)", "비고"])
    for r in df_in.values.tolist(): ws_in.append(r)
    in_max = ws_in.max_row
    if in_max > 1:
        ws_in.append(["합계", "", f"=SUM(C2:C{in_max})", "", f"=SUM(E2:E{in_max})", "", ""])
        for cell in ws_in[in_max + 1]:
            cell.font = sum_font; cell.fill = sum_fill; cell.border = sum_border; cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column in [3, 5]: cell.number_format = "#,##0"
    fill_in = PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid") 
    apply_log_sheet_style(ws_in, fill_in, Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF"), is_integrated=True)
    for col in ws_in.columns:
        max_len = max(sum(2 if ord(c) > 128 else 1 for c in str(cell.value or '')) for cell in col)
        ws_in.column_dimensions[col[0].column_letter].width = max(max_len + 4, 16)

    wb.remove(ws_default)
    wb.save(file_name)
    return file_name

# --- 메인 웹 앱 실행부 ---
init_db()

# 사이드바 구성
with st.sidebar:
    if os.path.exists("logo.webp"):
        try:
            image = Image.open("logo.webp")
            st.image(image, width=70)
        except Exception:
            st.title("🏬 JINTECH")
    else:
        st.title("🏬 JINTECH")
    
    st.markdown("**(주)진테크 스마트 관제**")
    menu = st.radio("메뉴 선택", ["🔄 입출고 관리", "📊 실시간 재고 모니터링", "👥 시스템 마스터 등록 관리"])
    st.markdown("---")
    st.caption("Designer. 수진")

# 탭 1: 입출고 관리
if menu == "🔄 입출고 관리":
    st.title("🔄 입출고 등록 및 로그 관리 대시보드")
    col1, col2 = st.columns([1, 1.8])
    
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM items ORDER BY name ASC")
    items_list = [r[0] for r in cursor.fetchall()]
    cursor.execute("SELECT name FROM workers ORDER BY name ASC")
    workers_list = [r[0] for r in cursor.fetchall()]
    conn.close()

    with col1:
        st.subheader("📝 데이터 장부 입력")
        d_date = st.date_input("날짜", datetime.date.today())
        
        selected_item = st.selectbox("품목 선택", items_list if items_list else ["등록된 품목 없음"])
        
        # 단가 실시간 표시
        item_price = 0
        if selected_item and selected_item != "등록된 품목 없음":
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            cursor.execute("SELECT price FROM items WHERE name=?", (selected_item,))
            r = cursor.fetchone()
            if r: item_price = r[0]
            conn.close()
        st.caption(f"💡 선택 품목 현재 단가: **{item_price:,}원**")

        division = st.radio("구분", ["출고", "입고"], horizontal=True)
        qty = st.number_input("수량", min_value=1, value=1, step=1)
        
        if division == "출고":
            manager = st.selectbox("작업자 선택 (가져가는 사람)", workers_list if workers_list else ["등록된 작업자 없음"])
        else:
            manager = st.text_input("입고처 (발주/주문 업체명)", placeholder="예: 대성자재")
            if not manager: manager = "주문업체 수령"

        note = st.text_input("비고 메모", placeholder="특이사항 기재")

        if st.button("📥 내역 등록 및 재고 반영", use_container_width=True, type="primary"):
            if not selected_item or selected_item == "등록된 품목 없음":
                st.error("품목을 먼저 등록 및 선택해 주세요.")
            elif division == "출고" and (not manager or manager == "등록된 작업자 없음"):
                st.error("작업자를 선택해 주세요.")
            else:
                conn = sqlite3.connect(DB_NAME)
                cursor = conn.cursor()
                cursor.execute("INSERT INTO history (date, name, division, qty, manager, note) VALUES (?, ?, ?, ?, ?, ?)",
                               (d_date.strftime("%Y-%m-%d"), selected_item, division, qty, manager, note))
                conn.commit()
                conn.close()
                st.success(f"[{selected_item}] {qty}개 {division} 등록 완료!")
                st.rerun()

    with col2:
        st.subheader("📊 요약 품목별 현재고 실시간 현황")
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT name, specification, safety_stock FROM items")
        all_items = cursor.fetchall()
        stock_dict = {name: {"spec": spec, "stock": 0} for name, spec, safety in all_items}
        cursor.execute("SELECT name, division, qty FROM history")
        for name, div, q in cursor.fetchall():
            if name in stock_dict:
                if div == "입고": stock_dict[name]["stock"] += q
                elif div == "출고": stock_dict[name]["stock"] -= q
        conn.close()
        
        df_stock_summary = pd.DataFrame([{"품목명": k, "규격/단위": v["spec"], "현재재고량": f"{v['stock']:,}"} for k, v in stock_dict.items()])
        st.dataframe(df_stock_summary, use_container_width=True, hide_index=True)

        st.subheader("📜 최근 입출고 추적 로그 (최근 50건)")
        conn = sqlite3.connect(DB_NAME)
        df_log = pd.read_sql_query("SELECT date AS '날짜', name AS '품목명', division AS '구분', qty AS '수량', manager AS '담당자/작업자', note AS '비고' FROM history ORDER BY id DESC LIMIT 50", conn)
        conn.close()
        st.dataframe(df_log, use_container_width=True, hide_index=True)

# 탭 2: 실시간 재고 모니터링
elif menu == "📊 실시간 재고 모니터링":
    st.title("📊 창고 내 스마트 통합 관제실")
    
    tab_mode = st.radio("관제 현황 선택", ["📦 현재고 종합 현황판", "📥 입고 내역 추적 장부", "📤 출고 내역 추적 장부"], horizontal=True)

    conn = sqlite3.connect(DB_NAME)
    if tab_mode == "📦 현재고 종합 현황판":
        cursor = conn.cursor()
        cursor.execute("SELECT name, specification, safety_stock, price FROM items")
        stock_dict = {name: {"spec": spec, "safety": safety, "price": price, "stock": 0} for name, spec, safety, price in cursor.fetchall()}
        cursor.execute("SELECT name, division, qty FROM history")
        for name, div, q in cursor.fetchall():
            if name in stock_dict:
                if div == "입고": stock_dict[name]["stock"] += q
                elif div == "출고": stock_dict[name]["stock"] -= q
        conn.close()

        data = []
        for k, v in stock_dict.items():
            status = "정상" if v["stock"] >= v["safety"] else "🚨 재고 부족"
            data.append({"품목명": k, "규격/단위": v["spec"], "현재고 수량": v["stock"], "안전재고": v["safety"], "등록단가(원)": f"{v['price']:,}", "상태": status})
        st.dataframe(pd.DataFrame(data), use_container_width=True, hide_index=True)

    elif tab_mode == "📥 입고 내역 추적 장부":
        query = """
            SELECT h.date AS '입고날짜', h.name AS '품목명', h.qty AS '수량', i.price AS '단가(원)', 
                   (h.qty * i.price) AS '입고금액(원)', h.manager AS '발주/입고처', h.note AS '비고' 
            FROM history h LEFT JOIN items i ON h.name = i.name WHERE h.division='입고' ORDER BY h.id DESC
        """
        df_in_view = pd.read_sql_query(query, conn)
        conn.close()
        st.dataframe(df_in_view, use_container_width=True, hide_index=True)

    elif tab_mode == "📤 출고 내역 추적 장부":
        query = """
            SELECT h.date AS '출고일자', h.name AS '품목명', h.qty AS '수량', i.price AS '단가(원)', 
                   (h.qty * i.price) AS '출고금액(원)', h.manager AS '작업자명단', h.note AS '비고' 
            FROM history h LEFT JOIN items i ON h.name = i.name WHERE h.division='출고' ORDER BY h.id DESC
        """
        df_out_view = pd.read_sql_query(query, conn)
        conn.close()
        st.dataframe(df_out_view, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.subheader("🗓️ 월간 종합 정산 및 멀티탭 결산 엑셀 보고서 출력 (자동 수식 연동)")
    
    col_s, col_e, col_btn = st.columns([1, 1, 1.5])
    today_now = datetime.date.today()
    first_day = today_now.replace(day=1)
    
    with col_s:
        s_date_in = st.date_input("조회 시작일", first_day)
    with col_e:
        e_date_in = st.date_input("조회 종료일", today_now)
    with col_btn:
        st.write(" ")
        st.write(" ")
        if st.button("📊 지정 기간 통합 결산보고서 엑셀 생성", use_container_width=True, type="primary"):
            file_created = generate_monthly_excel(s_date_in.strftime("%Y-%m-%d"), e_date_in.strftime("%Y-%m-%d"))
            if file_created and os.path.exists(file_created):
                with open(file_created, "rb") as fp:
                    st.download_button(
                        label="💾 생성된 엑셀 보고서 파일 즉시 다운로드",
                        data=fp,
                        file_name=file_created,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                st.success("결산보고서가 성공적으로 빌드되었습니다! 다운로드 버튼을 눌러주세요.")
            else:
                st.warning("해당 기간에는 조회할 수 있는 입출고 데이터가 없습니다.")

# 탭 3: 시스템 마스터 등록 관리
elif menu == "👥 시스템 마스터 등록 관리":
    st.title("👥 진테크 마스터 인프라 등록 관리 통제실")
    col_m1, col_m2 = st.columns(2)

    with col_m1:
        st.subheader("📦 1. 부자재 품목 마스터 정보 등록")
        m_name = st.text_input("품목명 (중복불가)", placeholder="예: 가스켓")
        m_spec = st.text_input("규격 / 단위", placeholder="예: 1ea 또는 40*40")
        m_safety = st.number_input("안전재고 기준량", min_value=0, value=0)
        m_price = st.number_input("개당 부자재 단가 (원)", min_value=0, value=0, step=100)

        if st.button("➕ 새 부자재 등록", use_container_width=True):
            if not m_name:
                st.error("품목명을 입력해 주세요.")
            else:
                conn = sqlite3.connect(DB_NAME)
                cursor = conn.cursor()
                try:
                    cursor.execute("INSERT INTO items (name, specification, safety_stock, price) VALUES (?, ?, ?, ?)",
                                   (m_name.strip(), m_spec.strip(), m_safety, m_price))
                    conn.commit()
                    st.success(f"[{m_name}] 등록 완료 (단가: {m_price:,}원)")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("이미 존재하는 품목명입니다.")
                finally:
                    conn.close()

        st.markdown("#### 등록된 품목 마스터 목록")
        conn = sqlite3.connect(DB_NAME)
        df_items = pd.read_sql_query("SELECT name AS '품목명', specification AS '규격/단위', safety_stock AS '안전재고', price AS '단가(원)' FROM items", conn)
        conn.close()
        st.dataframe(df_items, use_container_width=True, hide_index=True)

    with col_m2:
        st.subheader("👥 2. 작업자(직원) 마스터 등록")
        w_name = st.text_input("작업자 직원 이름 (중복불가)", placeholder="예: 김세령")

        if st.button("👥 새 작업자 등록", use_container_width=True):
            if not w_name:
                st.error("이름을 입력해 주세요.")
            else:
                conn = sqlite3.connect(DB_NAME)
                cursor = conn.cursor()
                try:
                    cursor.execute("INSERT INTO workers (name) VALUES (?)", (w_name.strip(),))
                    conn.commit()
                    st.success(f"[{w_name}] 작업자가 명단에 추가되었습니다.")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("이미 명단에 존재하는 이름입니다.")
                finally:
                    conn.close()

        st.markdown("#### 등록된 작업자 명단 목록")
        conn = sqlite3.connect(DB_NAME)
        df_workers = pd.read_sql_query("SELECT name AS '작업자 직원 명단' FROM workers ORDER BY name ASC", conn)
        conn.close()
        st.dataframe(df_workers, use_container_width=True, hide_index=True)